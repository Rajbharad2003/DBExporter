import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from db_connector import connect_db
from utils import clean_text, clean_table_name
import datetime

def get_tables(conn, db_type):
    cursor = conn.cursor()
    if db_type == "PostgreSQL":
        cursor.execute("""
            SELECT table_name FROM information_schema.tables 
            WHERE table_schema='public' AND table_type='BASE TABLE'
        """)
    elif db_type == "MySQL":
        cursor.execute("SHOW TABLES")
    return [row[0] for row in cursor.fetchall()]

def export_to_excel(config):
    try:
        conn = connect_db(config)
        cursor = conn.cursor()
        tables = get_tables(conn, config["db_type"])

        # if config["table_limit"] > 0:
        #     tables = tables[:config["table_limit"]]

        # if config["show_tables"]:
        #     print("📋 Tables:", tables)

        wb = Workbook()
        ws = wb.active
        ws.title = "Database Tables"

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for table in tables:
            query = f'SELECT * FROM `{table}` LIMIT {config["row_limit"]}' if config["db_type"] == "MySQL" \
                else f'SELECT * FROM "{table}" LIMIT {config["row_limit"]}'

            df = pd.read_sql_query(query, conn)

            if df.empty:
                continue

            row_start = ws.max_row + 2 if ws.max_row > 1 else 1
            ws.cell(row=row_start, column=1, value=f"Table: {clean_table_name(table)}").font = Font(bold=True, size=14)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=row_start + 2):
                for c_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=clean_text(val))
                    cell.border = border
                    if r_idx == row_start + 2:
                        cell.font = Font(bold=True)

        filename = f"Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        conn.close()
        return True, f"Exported successfully to '{filename}'"
    except Exception as e:
        return False, str(e)
